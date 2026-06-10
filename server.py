"""
CanSat Ground Station — WebSocket Server v2
============================================
รับข้อมูล CSV จาก ESP32 ผ่าน Serial (USB)
แล้ว broadcast ไปยัง browser dashboard ผ่าน WebSocket
+ Export ข้อมูลเป็น Excel (.xlsx) พร้อม MAX/MIN/AVG

Dependencies:
    pip install pyserial websockets openpyxl

Usage:
    python server.py --port COM3         (Windows)
    python server.py --port /dev/ttyUSB0 (Linux/Mac)
    python server.py --demo              (ทดสอบโดยไม่มี hardware)
    python server.py --demo --export     (demo แล้ว export excel อัตโนมัติ)
"""

import asyncio, json, argparse, logging, random, math, os, socket as _socket
from datetime import datetime

import serial
import serial.tools.list_ports
import websockets

# ─── openpyxl imports ───────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import DataPoint

# ─── Config ──────────────────────────────────────────────────────────────────
BAUD_RATE = 115200
WS_HOST   = "localhost"
WS_PORT   = 8765
LOG_CSV   = "telemetry_log.csv"

CSV_HEADERS = [
    "team_id","time","packet_id","lat","lon","sat",
    "temp","humidity","alt_baro",
    "acc_x","acc_y","acc_z","heading",
    "pm1_0","pm2_5","pm10",
    "voltage","current","watt","battery_percent","status"
]

NUMERIC_FIELDS = [
    "temp","humidity","alt_baro",
    "acc_x","acc_y","acc_z","heading",
    "pm1_0","pm2_5","pm10",
    "voltage","current","watt","battery_percent",
    "pitch","roll","tilt"
]

CSV_FILTER = set(CSV_HEADERS) | {"timestamp"}

SENSOR_PARSERS = {
    "BME": ["temp","humidity","alt_baro"],
    "GPS": ["lat","lon","sat"],
    "IMU": ["acc_x","acc_y","acc_z","heading"],
    "PM":  ["pm1_0","pm2_5","pm10"],
    "PWR": ["voltage","current"],
}

# ─── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger("cansat")

# ─── Globals ─────────────────────────────────────────────────────────────────
connected_clients:  set = set()
latest_packet:     dict = {}
packet_log:        list = []
session_excel_path: str = ""


# ─── Helpers ──────────────────────────────────────────────────────────────────
def parse_sensor_line(line: str):
    """Parse ESP32 single-sensor format: PREFIX:v1,v2,... → sensor dict or None"""
    line = line.strip()
    if ":" not in line:
        return None
    prefix, rest = line.split(":", 1)
    keys = SENSOR_PARSERS.get(prefix.upper())
    if not keys:
        return None
    parts = rest.split(",")
    if len(parts) < len(keys):
        return None
    try:
        data = {k: float(v) for k, v in zip(keys, parts)}
        data["type"]      = "sensor"
        data["sensor"]    = prefix.upper()
        data["timestamp"] = datetime.now().isoformat()
        return data
    except ValueError:
        return None


def parse_csv_line(line: str):
    line = line.strip()
    if not line or line.startswith("team_id"):
        return None
    parts = line.split(",")
    n = len(parts)
    EXPECTED = len(CSV_HEADERS)

    if n > EXPECTED:
        # Multiple LoRa packets concatenated (SF11 ToA > send interval).
        # Each boundary: status of record K and team_id of record K+1 are merged
        # e.g. status="1" + team_id="14" → "114" as one field.
        # Layout: R1 = parts[0:21] (21 fields, team_id present)
        #         R2 = [team_id] + parts[21:41] (20 fields, team_id stripped into R1's last)
        #         R3 = partial, skip
        team_id_val = parts[0].strip()
        n_ti = len(team_id_val)

        def _fix_last(chunk):
            chunk = list(chunk)
            last = chunk[-1].strip()
            if last.endswith(team_id_val) and len(last) > n_ti:
                try:
                    int(last[:-n_ti])
                    chunk[-1] = last[:-n_ti]
                except ValueError:
                    pass
            return chunk

        best = None
        # Prefer R2 (more recent) if complete
        if n >= EXPECTED + (EXPECTED - 1):
            r2 = _fix_last([team_id_val] + parts[EXPECTED: EXPECTED + (EXPECTED - 1)])
            if len(r2) == EXPECTED:
                best = r2
        # Fall back to R1
        if best is None:
            best = _fix_last(parts[:EXPECTED])

        if best:
            parts = best
            n = EXPECTED
        else:
            log.warning(f"Wrong field count ({n}): {line}")
            return None

    if n == EXPECTED - 1:
        parts.append("0")   # status field missing — default ascending
    elif n != EXPECTED:
        log.warning(f"Wrong field count ({n}): {line}")
        return None
    data = {}
    for k, v in zip(CSV_HEADERS, parts):
        if k == "team_id":
            data[k] = v.strip()
        else:
            try:
                data[k] = float(v)
            except (ValueError, TypeError):
                log.warning(f"Corrupt field '{k}': {repr(v)} — packet dropped")
                return None
    data["timestamp"] = datetime.now().isoformat()
    return data


def decode_status(status: int) -> dict:
    return {
        "ascending":  bool(status & 1),
        "apogee":     bool(status & 2),
        "deployment": bool(status & 4),
        "descending": bool(status & 8),
        "landed":     bool(status & 16),
    }


DROPOUT_ZERO_FIELDS    = {"alt_baro", "voltage", "temp", "humidity", "pm1_0", "pm2_5", "pm10"}
DROPOUT_NEGONE_FIELDS  = {"voltage", "current", "watt", "pm1_0", "pm2_5", "pm10", "battery_percent"}


# ─── Kalman Filter ────────────────────────────────────────────────────────────
class KalmanFilter1D:
    """Single-variable Kalman filter.
    q = process noise  (how fast the true value can change between packets)
    r = measurement noise (how noisy the sensor reading is)
    """
    def __init__(self, q: float = 0.1, r: float = 1.0):
        self.q = q
        self.r = r
        self.x = None        # current estimate (None = uninitialised)
        self.p: float = 1.0  # estimate error variance

    def update(self, z: float) -> float:
        if self.x is None:
            self.x = z
            return z
        self.p += self.q                     # predict: uncertainty grows
        k      = self.p / (self.p + self.r)  # Kalman gain
        self.x += k * (z - self.x)           # correct estimate
        self.p *= (1.0 - k)                  # shrink uncertainty
        return self.x

    def reset(self):
        self.x = None
        self.p = 1.0


KF_CONFIGS: dict = {
    # field          q (process noise)   r (sensor noise)
    "alt_baro":   (0.5,    2.0),   # barometer moderately noisy
    "temp":       (0.02,   0.5),   # temperature changes slowly
    "humidity":   (0.05,   1.0),
    "acc_x":      (80.0,  150.0),  # accelerometer very noisy
    "acc_y":      (80.0,  150.0),
    "acc_z":      (80.0,  150.0),
    "pm1_0":      (1.0,    5.0),
    "pm2_5":      (1.0,    5.0),
    "pm10":       (1.0,    5.0),
    "voltage":    (0.005,  0.05),  # voltage changes very slowly
}

kf_filters: dict = {f: KalmanFilter1D(q, r) for f, (q, r) in KF_CONFIGS.items()}


def apply_kalman(data: dict) -> None:
    """Apply each field's Kalman filter; writes result as <field>_kf in-place.
    Dropout sentinels (0 and -1) are skipped so they don't poison filter state.
    """
    for field, kf in kf_filters.items():
        val = data.get(field)
        if val is None:
            continue
        if field in DROPOUT_ZERO_FIELDS  and val == 0:
            continue
        if field in DROPOUT_NEGONE_FIELDS and val == -1:
            continue
        data[f"{field}_kf"] = round(kf.update(val), 3)


def reset_kalman() -> None:
    """Reset all filters at session start so history doesn't bleed across sessions."""
    for kf in kf_filters.values():
        kf.reset()


def compute_stats_for_packets(packets: list) -> dict:
    """คำนวณ MAX / MIN / AVG / STDDEV จาก list ของ packet"""
    if not packets:
        return {}
    stats = {}
    all_fields = NUMERIC_FIELDS + ["speed", "vacc"]
    for field in all_fields:
        vals = [p[field] for p in packets if field in p and p[field] is not None]
        if field in DROPOUT_ZERO_FIELDS:
            vals = [v for v in vals if v != 0]
        if field in DROPOUT_NEGONE_FIELDS:
            vals = [v for v in vals if v != -1]
        if vals:
            avg = sum(vals) / len(vals)
            stddev = math.sqrt(sum((v - avg) ** 2 for v in vals) / len(vals)) if len(vals) > 1 else 0
            stats[field] = {
                "max":    round(max(vals), 3),
                "min":    round(min(vals), 3),
                "avg":    round(avg, 3),
                "stddev": round(stddev, 3),
                "count":  len(vals),
            }
    return stats


def compute_stats() -> dict:
    """คำนวณ MAX / MIN / AVG / STDDEV จากข้อมูลทั้งหมดที่เก็บไว้"""
    return compute_stats_for_packets(packet_log)


TIME_GLITCH_THRESHOLD = 30  # วินาที — ถ้า time ถอยหลังเกินนี้ถือเป็น glitch/session ใหม่


def compute_derived(packets: list) -> list:
    """คำนวณ speed และ vacc จาก packet ที่ต่อเนื่องกัน
    Packets ที่ time กระโดดถอยหลังเกิน TIME_GLITCH_THRESHOLD จะถูก skip
    เพื่อกรอง firmware noise และตัด session ใหม่ที่ปนมา
    vacc ใช้ acc_y โดยตรง (แกน Y ชี้ขึ้น, รวม gravity ~980 mg)
    """
    R = 6371000
    result = []
    prev = None
    for pkt in packets:
        if prev and pkt["time"] < prev["time"] - TIME_GLITCH_THRESHOLD:
            log.warning(f"Time glitch: {prev['time']:.1f} → {pkt['time']:.1f} — packet skipped")
            continue
        d = dict(pkt)
        speed = 0.0
        if prev and d["time"] > prev["time"]:
            dt = d["time"] - prev["time"]
            if 0 < dt < 5:
                gps_valid = (prev["lat"] != -1 and prev["lon"] != -1 and
                             d["lat"]   != -1 and d["lon"]   != -1)
                if gps_valid:
                    dlat = (d["lat"] - prev["lat"]) * math.pi / 180
                    dlon = (d["lon"] - prev["lon"]) * math.pi / 180
                    a = math.sin(dlat/2)**2 + math.cos(prev["lat"]*math.pi/180) * math.cos(d["lat"]*math.pi/180) * math.sin(dlon/2)**2
                    dH = R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
                else:
                    dH = 0.0
                dV = 0.0 if d["alt_baro"] == 0 or prev["alt_baro"] == 0 else d["alt_baro"] - prev["alt_baro"]
                speed = max(0, math.sqrt(dH**2 + dV**2) / dt)
        d["speed"] = round(speed, 3)
        d["vacc"]  = round(d.get("acc_y") or 0.0, 3)
        ax = d.get("acc_x"); ay = d.get("acc_y"); az = d.get("acc_z")
        if None not in (ax, ay, az):
            mag = math.sqrt(ax*ax + ay*ay + az*az)
            if mag > 10:
                d["pitch"] = round(math.degrees(math.atan2(ay, math.sqrt(ax*ax + az*az))), 2)
                d["roll"]  = round(math.degrees(math.atan2(-ax, az)), 2)
                d["tilt"]  = round(math.degrees(math.atan2(math.sqrt(ax*ax + ay*ay), az)), 2)
            else:
                d["pitch"] = d["roll"] = d["tilt"] = None
        else:
            d["pitch"] = d["roll"] = d["tilt"] = None
        result.append(d)
        prev = pkt
    return result


# ─── Excel Export ─────────────────────────────────────────────────────────────
def export_excel(filename: str = None):
    if not packet_log:
        log.warning("No data to export")
        return None

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cansat_telemetry_{ts}.xlsx"

    wb = Workbook()

    # ── สีที่ใช้ ──────────────────────────────────────────────────────────────
    C_HEADER_BG  = "0D1B2A"   # dark navy
    C_HEADER_FG  = "00D4FF"   # cyan
    C_SUBHDR_BG  = "112233"
    C_MAX_BG     = "1A3A2A"
    C_MIN_BG     = "3A1A1A"
    C_AVG_BG     = "1A2A3A"
    C_ROW_ODD    = "0A1628"
    C_ROW_EVEN   = "061020"
    C_ACCENT     = "00FF88"
    C_BORDER     = "1A3A5A"

    thin = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(color=C_HEADER_FG, bold=True, size=11):
        return Font(name="Consolas", bold=bold, color=color, size=size)

    def cell_font(color="AADDFF", size=10):
        return Font(name="Consolas", color=color, size=size)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Raw Telemetry
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📡 Telemetry"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A4"

    # Title row
    ws1.merge_cells("A1:X1")
    title_cell = ws1["A1"]
    title_cell.value = f"🛰  CanSat Telemetry Log  —  {datetime.now().strftime('%d %b %Y  %H:%M:%S')}  —  {len(packet_log)} packets"
    title_cell.font      = Font(name="Consolas", bold=True, color=C_ACCENT, size=13)
    title_cell.fill      = fill(C_HEADER_BG)
    title_cell.alignment = center
    ws1.row_dimensions[1].height = 28

    # Derived data (with speed + vacc)
    derived_log = compute_derived(packet_log)

    # Inject wall_clock (HH:MM:SS) from ISO timestamp into each packet
    for pkt in derived_log:
        try:
            pkt["wall_clock"] = datetime.fromisoformat(pkt.get("timestamp","")).strftime("%H:%M:%S")
        except Exception:
            pkt["wall_clock"] = "--:--:--"

    # Field metadata (shared between Sheet 1 and Sheet 2)
    units = {
        "temp":"°C","humidity":"%","alt_baro":"m",
        "acc_x":"mg","acc_y":"mg","acc_z":"mg","heading":"°",
        "pm1_0":"µg/m³","pm2_5":"µg/m³","pm10":"µg/m³",
        "voltage":"V","current":"mA","watt":"W","battery_percent":"%",
        "speed":"m/s","vacc":"m/s²",
        "pitch":"°","roll":"°","tilt":"°",
    }
    labels = {
        "temp":"Temperature","humidity":"Humidity","alt_baro":"Altitude",
        "acc_x":"Acceleration X","acc_y":"Acceleration Y","acc_z":"Acceleration Z",
        "heading":"Heading","pm1_0":"PM 1.0","pm2_5":"PM 2.5","pm10":"PM 10",
        "voltage":"Voltage","current":"Current","watt":"Power","battery_percent":"Battery",
        "speed":"Speed (calc)","vacc":"Vert Accel (calc)",
        "pitch":"Pitch (calc)","roll":"Roll (calc)","tilt":"Tilt (calc)",
    }

    # Column headers — raw CSV keys first (row 2), display names (row 3)
    export_keys = ["team_id","time","wall_clock","packet_id","lat","lon","sat",
                   "temp","humidity","alt_baro","acc_x","acc_y","acc_z","heading",
                   "pm1_0","pm2_5","pm10","voltage","current","watt","battery_percent",
                   "status","speed","vacc","pitch","roll","tilt"]
    display_headers = [
        "Team ID","Time (s)","Clock","Packet ID","Latitude","Longitude","Satellites",
        "Temp (°C)","Humidity (%)","Altitude (m)",
        "Acc X (mg)","Acc Y (mg)","Acc Z (mg)","Heading (°)",
        "PM1.0 (µg)","PM2.5 (µg)","PM10 (µg)",
        "Voltage (V)","Current (mA)","Watt (W)","Battery (%)","Status",
        "Speed (m/s)","Vert Accel (m/s²)","Pitch (°)","Roll (°)","Tilt (°)"
    ]

    # Row 2: machine-readable keys (for replay)
    for col_idx, key in enumerate(export_keys, 1):
        cell = ws1.cell(row=2, column=col_idx, value=key)
        cell.font      = Font(name="Consolas", color="334455", size=8)
        cell.fill      = fill("060E1A")
        cell.alignment = center
        cell.border    = border
    ws1.row_dimensions[2].height = 14

    # Row 3: display headers
    for col_idx, header in enumerate(display_headers, 1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font      = hdr_font()
        cell.fill      = fill(C_SUBHDR_BG)
        cell.alignment = center
        cell.border    = border
    ws1.row_dimensions[3].height = 22

    # Data rows (start row 4)
    for row_idx, pkt in enumerate(derived_log, 4):
        session_idx = row_idx - 3  # 1-based counter since START
        bg = C_ROW_ODD if row_idx % 2 == 0 else C_ROW_EVEN
        for col_idx, key in enumerate(export_keys, 1):
            val  = pkt.get(key, "")
            if key == "time":
                display_val = f"{round(val, 1) if isinstance(val, float) else val} ({session_idx})"
            elif key == "packet_id":
                display_val = f"{int(val) if isinstance(val, (float, int)) else val} ({session_idx})"
            else:
                display_val = round(val, 4) if isinstance(val, float) else val
            cell = ws1.cell(row=row_idx, column=col_idx, value=display_val)
            cell.font      = cell_font()
            cell.fill      = fill(bg)
            cell.alignment = center
            cell.border    = border
        ws1.row_dimensions[row_idx].height = 16

    # Column widths
    widths = [10,9,8,9,12,12,10, 10,11,11, 10,10,10,10, 10,10,10, 10,11,10,10,8, 10,12, 9,9,9]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Statistics by Flight Phase + Charts
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("📊 Statistics")
    ws2.sheet_view.showGridLines = False

    team = packet_log[0].get("team_id","—") if packet_log else "—"
    start_clock = derived_log[0].get("wall_clock","—") if derived_log else "—"
    end_clock   = derived_log[-1].get("wall_clock","—") if derived_log else "—"
    duration = 0
    if len(derived_log) >= 2:
        try:
            t0 = datetime.strptime(start_clock, "%H:%M:%S")
            t1 = datetime.strptime(end_clock,   "%H:%M:%S")
            duration = round((t1 - t0).total_seconds(), 1)
        except Exception:
            pass

    ws2.merge_cells("A1:L1")
    c = ws2["A1"]
    c.value = f"📊  Statistics by Phase  —  Team: {team}  |  {len(derived_log)} packets  |  {start_clock} – {end_clock}  |  {duration}s"
    c.font      = Font(name="Consolas", bold=True, color=C_ACCENT, size=13)
    c.fill      = fill(C_HEADER_BG)
    c.alignment = center
    ws2.row_dimensions[1].height = 28

    PHASE_GROUPS = [
        (1,  "ASCENDING",  "1A3A2A", "00FF88", "🚀 ASCENDING"),
        (2,  "APOGEE",     "102030", "00D4FF", "🎯 APOGEE"),
        (4,  "DEPLOYMENT", "2A2010", "FFC040", "🪂 DEPLOYMENT"),
        (8,  "DESCENDING", "1A1030", "BB88FF", "⬇ DESCENDING"),
        (16, "LANDED",     "3A1020", "FF6688", "🏁 LANDED"),
    ]

    # ── Collect per-phase stats (only phases with data) ──────────────────
    active_phases = []
    for mask, name, pbg, pfg, ename in PHASE_GROUPS:
        pkts = [p for p in derived_log if int(p.get("status", 0)) & mask]
        if pkts:
            active_phases.append((name, compute_stats_for_packets(pkts), pbg, pfg, ename, len(pkts)))
    n_phases = len(active_phases)

    # ── Key metrics for summary charts ───────────────────────────────────
    CHART_METRICS = [
        ("alt_baro",        "Altitude",     "m"),
        ("temp",            "Temperature",  "°C"),
        ("humidity",        "Humidity",     "%"),
        ("battery_percent", "Battery",      "%"),
        ("pm2_5",           "PM 2.5",       "µg/m³"),
        ("voltage",         "Voltage",      "V"),
    ]
    STAT_ROWS = [
        ("MAX", "max", "00FF88", C_MAX_BG),
        ("AVG", "avg", "00D4FF", C_AVG_BG),
        ("MIN", "min", "FF6688", C_MIN_BG),
    ]
    PHASE_COL0 = 2  # phase data starts at column B (index 2)

    # ── Write compact summary tables + collect chart reference positions ──
    chart_refs = []  # (field_label, field_unit, hrow, drow0, drow_end)
    cur_row = 3

    for field, field_label, field_unit in CHART_METRICS:
        hrow  = cur_row
        drow0 = cur_row + 1

        # Header row: metric label | phase1 | phase2 | ...
        c = ws2.cell(row=cur_row, column=1, value=field_label)
        c.font = hdr_font(); c.fill = fill(C_SUBHDR_BG)
        c.alignment = center; c.border = border

        for ci, (name, stats, pbg, pfg, ename, _) in enumerate(active_phases):
            c = ws2.cell(row=cur_row, column=PHASE_COL0 + ci, value=name)
            c.font = hdr_font(color=pfg); c.fill = fill(pbg)
            c.alignment = center; c.border = border

        ws2.row_dimensions[cur_row].height = 18
        cur_row += 1

        # Data rows: MAX / AVG / MIN
        for stat_label, stat_key, stat_fg, stat_bg in STAT_ROWS:
            c = ws2.cell(row=cur_row, column=1, value=stat_label)
            c.font = Font(name="Consolas", bold=True, color=stat_fg, size=10)
            c.fill = fill(stat_bg); c.alignment = center; c.border = border

            for ci, (name, stats, pbg, pfg, ename, _) in enumerate(active_phases):
                val = stats.get(field, {}).get(stat_key)
                c = ws2.cell(row=cur_row, column=PHASE_COL0 + ci, value=val)
                c.font = Font(name="Consolas", color=stat_fg, size=10)
                c.fill = fill(C_ROW_ODD if ci % 2 == 0 else C_ROW_EVEN)
                c.alignment = center; c.border = border

            ws2.row_dimensions[cur_row].height = 16
            cur_row += 1

        chart_refs.append((field_label, field_unit, hrow, drow0, cur_row - 1))
        cur_row += 1  # blank separator between metrics

    # ── BarCharts (placed to the right of the summary tables) ────────────
    if n_phases > 0:
        chart_col = get_column_letter(PHASE_COL0 + n_phases + 2)

        for fi, (field_label, field_unit, hrow, drow0, drow_end) in enumerate(chart_refs):
            ch = BarChart()
            ch.type = "col"
            ch.grouping = "clustered"
            ch.title = f"{field_label} by Phase"
            ch.y_axis.title = f"{field_label} ({field_unit})"
            ch.x_axis.title = "Phase"
            ch.width = 16; ch.height = 10; ch.style = 10

            # Data: col A (stat labels = series names) + phase cols
            data_ref = Reference(ws2,
                                 min_col=1,
                                 max_col=PHASE_COL0 + n_phases - 1,
                                 min_row=drow0, max_row=drow_end)
            ch.add_data(data_ref, from_rows=True, titles_from_data=True)

            # Categories: phase names from header row
            cats_ref = Reference(ws2,
                                 min_col=PHASE_COL0,
                                 max_col=PHASE_COL0 + n_phases - 1,
                                 min_row=hrow, max_row=hrow)
            ch.set_categories(cats_ref)

            ws2.add_chart(ch, f"{chart_col}{3 + fi * 23}")

    # ── Column widths for summary area ───────────────────────────────────
    ws2.column_dimensions["A"].width = 16
    for ci in range(n_phases):
        ws2.column_dimensions[get_column_letter(PHASE_COL0 + ci)].width = 14

    # ── Separator before detailed tables ─────────────────────────────────
    cur_row += 1
    sep_col = get_column_letter(max(PHASE_COL0 + n_phases - 1, 7))
    ws2.merge_cells(f"A{cur_row}:{sep_col}{cur_row}")
    c = ws2.cell(row=cur_row, column=1, value="── Detailed Statistics per Phase ──")
    c.font = Font(name="Consolas", bold=True, color="556677", size=10)
    c.fill = fill(C_HEADER_BG); c.alignment = center
    ws2.row_dimensions[cur_row].height = 20
    cur_row += 1

    # ── Detailed per-phase tables (all metrics with STDDEV + count) ───────
    stat_fields = NUMERIC_FIELDS + ["speed", "vacc"]

    for mask, name, phase_bg, phase_fg, emoji_name in PHASE_GROUPS:
        phase_pkts = [p for p in derived_log if int(p.get("status", 0)) & mask]
        if not phase_pkts:
            continue

        ws2.merge_cells(f"A{cur_row}:G{cur_row}")
        c = ws2.cell(row=cur_row, column=1)
        c.value     = f"{emoji_name}  ·  {len(phase_pkts)} packets"
        c.font      = Font(name="Consolas", bold=True, color=phase_fg, size=11)
        c.fill      = fill(phase_bg); c.alignment = center
        ws2.row_dimensions[cur_row].height = 22
        cur_row += 1

        for col, txt in enumerate(["Parameter","MAX","MIN","AVG","STDDEV","Count","Unit"], 1):
            c = ws2.cell(row=cur_row, column=col, value=txt)
            c.font = hdr_font(); c.fill = fill(C_SUBHDR_BG)
            c.alignment = center; c.border = border
        ws2.row_dimensions[cur_row].height = 18
        cur_row += 1

        phase_stats = compute_stats_for_packets(phase_pkts)

        for r_i, field in enumerate(stat_fields):
            s  = phase_stats.get(field, {})
            bg = C_ROW_ODD if r_i % 2 == 0 else C_ROW_EVEN

            c = ws2.cell(row=cur_row, column=1, value=labels.get(field, field))
            c.font = hdr_font(color="AADDFF", bold=False)
            c.fill = fill(bg); c.alignment = left; c.border = border

            c = ws2.cell(row=cur_row, column=2, value=s.get("max","—"))
            c.font = Font(name="Consolas", color="00FF88", size=11, bold=True)
            c.fill = fill(C_MAX_BG); c.alignment = center; c.border = border

            c = ws2.cell(row=cur_row, column=3, value=s.get("min","—"))
            c.font = Font(name="Consolas", color="FF6688", size=11, bold=True)
            c.fill = fill(C_MIN_BG); c.alignment = center; c.border = border

            c = ws2.cell(row=cur_row, column=4, value=s.get("avg","—"))
            c.font = Font(name="Consolas", color="00D4FF", size=11, bold=True)
            c.fill = fill(C_AVG_BG); c.alignment = center; c.border = border

            c = ws2.cell(row=cur_row, column=5, value=s.get("stddev","—"))
            c.font = Font(name="Consolas", color="BB88FF", size=11)
            c.fill = fill(bg); c.alignment = center; c.border = border

            c = ws2.cell(row=cur_row, column=6, value=s.get("count", 0))
            c.font = cell_font(color="667788")
            c.fill = fill(bg); c.alignment = center; c.border = border

            c = ws2.cell(row=cur_row, column=7, value=units.get(field,""))
            c.font = cell_font(color="667788"); c.fill = fill(bg)
            c.alignment = center; c.border = border

            ws2.row_dimensions[cur_row].height = 18
            cur_row += 1

        cur_row += 1

    for col, w in zip("ABCDEFG", [24, 14, 14, 14, 14, 8, 12]):
        ws2.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Charts data (Altitude + Temp)
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("📈 Charts")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value = "📈  Mission Charts"
    c.font = Font(name="Consolas", bold=True, color=C_ACCENT, size=13)
    c.fill = fill(C_HEADER_BG); c.alignment = center
    ws3.row_dimensions[1].height = 28

    chart_keys = ["time","alt_baro","speed","temp","humidity","voltage","pm2_5"]
    chart_hdrs = ["Time (s)","Altitude (m)","Speed (m/s)","Temp (°C)","Humidity (%)","Voltage (V)","PM2.5 (µg)"]
    for col, hdr in enumerate(chart_hdrs, 1):
        c = ws3.cell(row=2, column=col, value=hdr)
        c.font = hdr_font(); c.fill = fill(C_SUBHDR_BG)
        c.alignment = center; c.border = border

    derived = compute_derived(packet_log)
    for i, pkt in enumerate(derived, 3):
        for col, key in enumerate(chart_keys, 1):
            c = ws3.cell(row=i, column=col, value=round(pkt.get(key,0), 3))
            c.font = cell_font()
            c.fill = fill(C_ROW_ODD if i%2==0 else C_ROW_EVEN)
            c.alignment = center; c.border = border

    n = len(packet_log)

    def make_chart(title, y_label, col_idx, color, anchor, w=22, h=12):
        ch = LineChart()
        ch.title = title; ch.style = 10
        ch.y_axis.title = y_label; ch.x_axis.title = "Time (s)"
        ch.height = h; ch.width = w
        dr = Reference(ws3, min_col=col_idx, min_row=2, max_row=n+2)
        cats = Reference(ws3, min_col=1, min_row=3, max_row=n+2)
        ch.add_data(dr, titles_from_data=True)
        ch.set_categories(cats)
        if ch.series:
            ch.series[0].graphicalProperties.line.solidFill = color
            ch.series[0].graphicalProperties.line.width = 20000
        ws3.add_chart(ch, anchor)

    if n > 1:
        make_chart("Altitude Profile",   "m",   2, "00D4FF", "I2")
        make_chart("Speed (calculated)", "m/s", 3, "00FFCC", "I24")
        make_chart("Temperature",        "°C",  4, "00FF88", "I46")
        make_chart("Humidity",           "%",   5, "FFC040", "I68")
        make_chart("Battery Voltage",    "V",   6, "FFC040", "I90")
        make_chart("PM 2.5",             "µg",  7, "0099CC", "I112")

    for col, w in zip("ABCDEFG", [10,12,10,10,11,10,10]):
        ws3.column_dimensions[col].width = w

    try:
        wb.save(filename)
    except PermissionError:
        # ไฟล์ถูกเปิดอยู่ใน Excel → บันทึกชื่อใหม่แทน
        alt = filename.replace(".xlsx", f"_export_{datetime.now().strftime('%H%M%S')}.xlsx")
        wb.save(alt)
        filename = alt
        log.warning(f"⚠️  PermissionError — saved to {alt} instead")
    log.info(f"✅ Excel exported → {filename}  ({len(packet_log)} rows)")
    return filename


# ─── Broadcast ────────────────────────────────────────────────────────────────
async def broadcast(message: dict):
    if not connected_clients:
        return
    payload = json.dumps(message)
    await asyncio.gather(
        *[c.send(payload) for c in connected_clients],
        return_exceptions=True
    )


# ─── WebSocket handler ────────────────────────────────────────────────────────
async def ws_handler(websocket):
    global latest_packet, session_excel_path
    connected_clients.add(websocket)
    log.info(f"Client connected ({len(connected_clients)} total)")

    if latest_packet:
        await websocket.send(json.dumps(latest_packet))
    if packet_log:
        await websocket.send(json.dumps({"type":"history","data":packet_log[-50:]}))

    try:
        async for msg in websocket:
            data = json.loads(msg)
            if data.get("cmd") == "start_session":
                packet_log.clear()
                latest_packet = {}
                reset_kalman()
                now = datetime.now()
                ts_label = now.strftime("%H:%M:%S")
                ts_file  = now.strftime("%Y%m%d_%H%M%S")
                base_dir = os.path.dirname(os.path.abspath(__file__))
                session_excel_path = os.path.join(base_dir, f"cansat_telemetry_{ts_file}.xlsx")
                # สร้างไฟล์ placeholder ทันทีที่กด START
                _wb = Workbook()
                _ws = _wb.active
                _ws.title = "📡 Telemetry"
                _ws["A1"] = f"Session started {ts_label} — recording…"
                _wb.save(session_excel_path)
                log.info(f"New session started at {ts_label} — Excel: {session_excel_path}")
                await websocket.send(json.dumps({
                    "type":  "session_started",
                    "time":  ts_label,
                    "excel": session_excel_path,
                }))
            elif data.get("cmd") == "stop_session":
                log.info(f"STOP — packet_log has {len(packet_log)} packets")
                try:
                    fname = export_excel(session_excel_path if session_excel_path else None)
                    session_excel_path = ""
                    stats = compute_stats()
                    await websocket.send(json.dumps({
                        "type":     "export_done",
                        "filename": fname,
                        "rows":     len(packet_log),
                        "stats":    stats,
                    }))
                except Exception as e:
                    log.error(f"stop_session error: {e}", exc_info=True)
                    await websocket.send(json.dumps({"type": "export_error", "error": str(e)}))
            elif data.get("cmd") == "export_excel":
                fname = export_excel()
                await websocket.send(json.dumps({
                    "type":     "export_done",
                    "filename": fname,
                    "rows":     len(packet_log),
                    "stats":    compute_stats()
                }))
            elif data.get("cmd") == "get_stats":
                await websocket.send(json.dumps({
                    "type":  "stats",
                    "stats": compute_stats(),
                    "total": len(packet_log)
                }))
            elif data.get("cmd") == "inject_packet":
                pkt = data.get("packet", {})
                pkt["type"] = "telemetry"
                pkt["timestamp"] = datetime.now().isoformat()
                apply_kalman(pkt)
                pkt["status_flags"] = decode_status(int(pkt.get("status", 0)))
                latest_packet = pkt
                packet_log.append({k: v for k, v in pkt.items() if k in CSV_FILTER})
                log.info(f"INJECT PKT#{int(pkt.get('packet_id',0)):04d}  alt={pkt.get('alt_baro',0):.0f}m  temp={pkt.get('temp',0):.1f}°C")
                await broadcast(pkt)
                await websocket.send(json.dumps({
                    "type":       "inject_ack",
                    "packet_id":  pkt.get("packet_id", 0),
                }))
    except websockets.exceptions.ConnectionClosed:
        pass
    finally:
        connected_clients.discard(websocket)
        log.info(f"Client disconnected ({len(connected_clients)} total)")


# ─── Serial reader (with auto-reconnect) ─────────────────────────────────────
async def serial_reader(port: str):
    global latest_packet
    loop = asyncio.get_event_loop()
    ser = None
    RETRY_DELAY = 3  # วินาที รอก่อน reconnect

    while True:
        # ── Connect ───────────────────────────────────────────────────────────
        try:
            ser = serial.Serial(port, BAUD_RATE, timeout=1)
            log.info(f"Serial open ✓  {port} @ {BAUD_RATE}")
            await broadcast({"type": "serial_status", "connected": True, "port": port})
        except serial.SerialException as e:
            log.warning(f"Cannot open {port}: {e} — retry in {RETRY_DELAY}s")
            await broadcast({"type": "serial_status", "connected": False, "port": port})
            await asyncio.sleep(RETRY_DELAY)
            continue

        # ── Read loop ─────────────────────────────────────────────────────────
        try:
            while True:
                line = await loop.run_in_executor(None, ser.readline)
                line = line.decode("utf-8", errors="ignore")
                if not line.strip():
                    await asyncio.sleep(0.01)
                    continue

                stripped = line.strip()
                if stripped.startswith("#"):
                    log.info(f"GS: {stripped}")
                    continue
                # Skip ESP32 boot messages (don't start with digit)
                if not stripped or not stripped[0].isdigit():
                    continue
                log.info(f"RAW: {stripped}")
                data = parse_sensor_line(line)
                if data is not None:
                    log.info(f"SENSOR {data['sensor']}: {line.strip()[:60]}")
                    await broadcast(data)
                    continue

                data = parse_csv_line(line)
                if data is None:
                    continue

                apply_kalman(data)
                data["status_flags"] = decode_status(int(data.get("status", 0)))
                latest_packet = {"type": "telemetry", **data}
                packet_log.append(data)

                log.info(f"PKT#{int(data['packet_id']):04d}  alt={data['alt_baro']:.0f}m  temp={data['temp']:.1f}°C  bat={data['battery_percent']:.0f}%")
                await broadcast(latest_packet)

        except serial.SerialException as e:
            log.warning(f"Serial disconnected: {e} — retry in {RETRY_DELAY}s")
            await broadcast({"type": "serial_status", "connected": False, "port": port})
        finally:
            try:
                ser.close()
            except Exception:
                pass
            ser = None

        await asyncio.sleep(RETRY_DELAY)


# ─── Demo generator ───────────────────────────────────────────────────────────
async def demo_generator(auto_export=False):
    global latest_packet
    log.info("DEMO MODE — simulated CanSat flight")
    t, pkt_id = 0.0, 1
    APOGEE, LAND = 20, 55

    while True:
        if t < APOGEE:
            alt    = (t / APOGEE) * 800
            status = 1
        elif t < LAND:
            frac   = (t - APOGEE) / (LAND - APOGEE)
            alt    = 800 * (1 - frac)
            status = 8 if frac > 0.1 else 2
        else:
            alt    = max(0, 10 - (t - LAND) * 5)
            status = 16

        data = {
            "type":"telemetry","timestamp":datetime.now().isoformat(),
            "time":round(t,1),"packet_id":pkt_id,
            "lat":round(18.12345 + t*0.00001, 6),"lon":round(98.12345 + t*0.000005, 6),
            "sat":10+random.randint(0,4),
            "temp":round(27 + t*0.05 + random.uniform(-0.3,0.3), 2),
            "humidity":round(65 + math.sin(t*0.1)*5, 1),
            "alt_baro":round(alt + random.uniform(-2,2), 1),
            "acc_x":round(random.uniform(-200,200),1),
            "acc_y":round(random.uniform(-200,200),1),
            "acc_z":round(980 + random.uniform(-50,50),1),
            "heading":round((t*3)%360, 1),
            "pm1_0":round(12+random.uniform(-2,5),1),
            "pm2_5":round(25+random.uniform(-3,8),1),
            "pm10":round(40+random.uniform(-5,10),1),
            "voltage":round(3.7-t*0.005, 2),
            "current":round(120+random.uniform(-10,10),1),
            "watt":round((3.7-t*0.005)*(120+random.uniform(-10,10))/1000, 3),
            "battery_percent":round(max(0,85-t*0.3),1),
            "status":status,
            "status_flags":decode_status(status),
        }

        latest_packet = data
        packet_log.append({k:v for k,v in data.items() if k in CSV_HEADERS+["timestamp"]})
        log.info(f"DEMO PKT#{pkt_id:04d}  alt={data['alt_baro']:.0f}m  status={status}")
        await broadcast(data)

        t += 0.5; pkt_id += 1
        await asyncio.sleep(0.5)

        if status == 16 and t > LAND + 5:
            log.info("Mission complete!")
            if auto_export:
                export_excel()
            await asyncio.sleep(5)
            t, pkt_id = 0.0, 1
            # packet_log ไม่ clear — ปล่อยให้ START button เป็นคนล้าง
            log.info("Simulation reset")


# ─── Main ─────────────────────────────────────────────────────────────────────
async def main(args):
    log.info("=" * 50)
    log.info("  CanSat Ground Station v2")
    log.info(f"  ws://{WS_HOST}:{WS_PORT}")
    log.info("=" * 50)

    sock = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
    # Windows SO_REUSEADDR lets multiple processes listen on the same port,
    # causing split traffic and 1011 errors. Use exclusive-address instead.
    if hasattr(_socket, "SO_EXCLUSIVEADDRUSE"):
        sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_EXCLUSIVEADDRUSE, 1)
    else:
        sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEADDR, 1)
    try:
        sock.bind(("127.0.0.1", WS_PORT))
    except OSError:
        log.error(f"Port {WS_PORT} is already in use — is another server.py running?")
        log.error("Run:  taskkill /F /IM python.exe   then restart.")
        sock.close()
        return
    ws_server = await websockets.serve(ws_handler, sock=sock)

    if args.demo:
        await asyncio.gather(ws_server.wait_closed(), demo_generator(args.export))
    else:
        await asyncio.gather(ws_server.wait_closed(), serial_reader(args.port))


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--port",   type=str, help="Serial port")
    p.add_argument("--baud",   type=int, default=115200)
    p.add_argument("--demo",   action="store_true")
    p.add_argument("--export", action="store_true", help="Auto-export Excel เมื่อ landed")
    p.add_argument("--list",   action="store_true")
    args = p.parse_args()

    if args.list:
        for p in serial.tools.list_ports.comports():
            print(f"  {p.device:20s} {p.description}")
    elif not args.demo and not args.port:
        print("Usage: python server.py --demo\n       python server.py --port COM3")
    else:
        BAUD_RATE = args.baud
        asyncio.run(main(args))
